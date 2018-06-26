import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import threading
import os
import utility
import DB
from bs4 import BeautifulSoup as bs
import pickle


class Excel_bot():
    def __init__(self, database, file=False, assume=False, unique='A'):
        '''

        '''
        self.region = pickle.load(open('r_key', 'rb'))
        self.cc = ''
        self.set = set()
        self.db = {}
        self.ucolumn = 0
        self.unique = unique
        self.a = assume
        self.oj = PatternFill(fill_type='solid',
                     start_color='f3af84',
                     end_color='f3af84')

        self.bord =  Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))


        self.og = PatternFill(fill_type='solid',
                     start_color='a9d08f',
                     end_color='a9d08f')

        self.align = Alignment(horizontal='center',
                               vertical='center',
                               text_rotation=0,
                               wrap_text=False,
                               shrink_to_fit=False,
                               indent=0)
        self.font = Font(name='Arial',
                         size=10,
                         bold=False,
                         italic=False,
                         vertAlign=None,
                         underline='none',
                         strike=False,
                         color='00000000')

        print("Loading database")
        self.dbm = DB.DBM(database, self)
        print("Loading file")
        if file:
            self.filepath = file
            if os.path.isfile(self.filepath):
                self.wb = openpyxl.load_workbook(self.filepath)
            else:
                self.wb = openpyxl.Workbook()
        else:
            self.filepath = None
            self.wb = openpyxl.Workbook()
        for sheet in self.wb.sheetnames:
            self.sheetn = sheet
            self.sheet = self.wb[self.sheetn]
            self.db[self.sheetn] = {"Meta": {"sheet name": self.sheetn,
                                             'width': self._current_width(),
                                             'bottom': self._current_line()}}
            self._sheet_load()
            self.db[sheet]['Data'] = self.self_scan()
        print("Loading complete")
        print("Choose an active sheet")

    def _sheet_load(self):
        '''
        Loads sheets into dictionary db
        '''
        ncolumns = self._column_scan()
        self.db[self.sheetn]["Meta"]["column names"] = ncolumns
        cd = {get_column_letter(x + 1): ncolumns[x]
              for x in range(len(ncolumns))}
        cdr = dict(map(reversed, cd.items()))
        scanlist = self.scan_list(self.unique)
        uidd = {x[0]: x[1] for x in scanlist}
        self.db[self.sheetn]["Meta"]["column dic"] = cd
        self.db[self.sheetn]["Meta"]["column dic r"] = cdr
        self.db[self.sheetn]["Meta"]["uid key dic"] = uidd

    def initialize(self):
        print('Running initial test:')
        with DB.DB(self.dbm, self.dbm.db):
            self.companies = set()
            self.rows = ['companies name', 'categories name', 'title', 'active_jobs ID',
                         'post_date', 'Location', 'description', 'qualifications']
            insert = '''
                    SELECT compn, catn, r.title, r.ID, dt, town, state, country_code, r.desc, r.p_desc
                    FROM    (
                            SELECT compn, catn, q.title, q.ID, datetime(post_date, 'unixepoch') as dt, q.desc, q.p_desc
                            FROM    (
                                    SELECT compn, catn, d.title, g.ID, d.desc, d.p_desc
                                    FROM    (
                                            SELECT c.name as compn, cat.name as catn, c.ID
                                            FROM    (
                                                    SELECT c.name, c.ID
                                                    FROM    (
                                                            SELECT * 
                                                            FROM company_ID
                                                            NATURAL JOIN 
                                                            companies
                                                            ) c
                                            LEFT JOIN active_jobs a on a.ID = c.ID
                                                    ) c
                                            LEFT JOIN       (
                                                            SELECT * 
                                                            FROM job_categories
                                                            NATURAL JOIN 
                                                            categories
                                                            ) cat on cat.ID = c.ID
                                            ) g
                                            LEFT JOIN desc as d on d.ID = g.ID
                                    ) q
                                    LEFT JOIN posting_date p on p.ID = q.id
                            ) r
                    LEFT JOIN location l on l.ID = r.ID
                    WHERE dt between '2018-01-01' AND '2018-04-30'
                    ''' #where dt between '2018-01-01' AND '2018-02-30'
            self.dbm.c.execute(insert)
            full = [x for x in self.dbm.c]
            # Nick edits
        full = [list(job[:5]) + [(job[7] if job[7] else '') + ', ' + (job[6] if job[6] else '')] + list(job[8:])
                for job in full]
        self.rows = [x.replace('_', ' ') for x in self.rows]
        while True:
            print("Method 2 is Adele's method")
            x = input('method1 or method2 (1/2): ')
            if x.isdigit():
                if int(x) == 1 or int(x) == 2:
                    break
            print('sorry this is an invalid input please try again')
        [self.method1, self.method2][int(x)-1](full)

    def alpha_f(self, row):
        return row[1:4]+[row[4].replace('-', ' ')]+[row[4][:-8].split('-')[y] for y in range(3)] + [row[4].split(' ')[0]]

    def beta_f(self, row):
        if 'NO' in row[5]:
            row[5] = row[5][4:]
        if ', ' in row[5]:
            split = row[5].split(', ')
            l = len(split)
            cc = split[0]
            if l >= 2:
                try:
                    state = split[1]
                except:
                    state = ''
        else:
            cc, state = row[5], ''
        cc = cc.upper()
        if len(cc) == 2:
            try:
                region = [self.region[cc]]
            except KeyError:
                x = list(set(self.region.values()))
                print(x)
                region = [x[int(input('Which region is %s ?'%cc))]]
                self.region[cc] = region[0]
                pickle.dump(self.region, open('r_key', 'wb'))
        else:
            region = ['Unknown']
        return [row[5]]+[state, cc.upper()]+region+[row[0]]

    def method2(self, full):
        lets = [chr(65 + x) for x in range(13)]
        widths = [3.63, 13.58, 3.26, 1.75, 0.76, 0.76, 0.76, 1.06, 8.03, 4.53, 1.25, 1.92, 1.19]       
        tits = ['categories name:','title:','active_jobs ID:','post_date:','YR','MM','DT','Date','Location (Original)','Location:','Country:','Buisness Region:','Company:']

        insert = [self.alpha_f(row)+self.beta_f(row) for row in full]
        for x in range(13):
            self.sheet['%s1'%lets[x]] = tits[x]
            self.sheet['%s1'%lets[x]].alignment = self.align
            self.sheet.column_dimensions[lets[x]].width = widths[x] * 10
            if 8> x >= 5 or 10<= x <=12:
                self.sheet['%s1'%lets[x]].fill = self.og
            else:
                self.sheet['%s1'%lets[x]].fill = self.oj
            self.sheet['%s1'%lets[x]].border = self.bord


        # indexing
        for x in range(1, len(full)):
            for y in range(len(insert[x])):
                index = '%s%s'%(get_column_letter(y+1), x+1)
                try:
                    if insert[x][y]:
                        self.sheet[index] = (insert[x][y].encode("utf8","backslashreplace")).decode("utf8","backslashreplace")
                    else:
                        self.sheet[index] = ''
                except openpyxl.utils.exceptions.IllegalCharacterError as ex:
                    self.sheet[index] = bs(insert[x][y], 'lxml').text
                if y >= 5 and y != 10:
                    self.sheet[index].alignment = self.align
                if 8> y >= 5 or 10 <= y <=12:
                    self.sheet[index].fill = self.og

                self.sheet[index].border = self.bord

        self.commit_to_xlxs()

    def method1(self, full):
        column_width = (max([len(x[1]) for x in full if x[1]]) + 2) * 1.2
        g = 0
        y = 0 
        for x in range(len(full)):
            if full[x][0] != self.cc:
                g = x
            self.compcheck(full[x][0])
            try:
                self.sheet.column_dimensions[get_column_letter(
                    (x + 2)-y)].width = column_width
            except ValueError as ex:
                self.compcheck(full[x][0]+'2')
                y = x
                self.sheet.column_dimensions[get_column_letter(
                    (x + 2)-y)].width = column_width
            if full[x][0] != self.cc:
                self.cc = full[x][0]
            self._outtosheet(x + 2 - g, full[x][1:], full[x][0])

        self.commit_to_xlxs()

    def sanatize(self, pack):
        if pack:
            if 'None' in str(pack):
                pack = pack.replace('None', '')
            if 'DESCRIPTION' in str(pack):
                pack = pack.replace('DESCRIPTION\n', '')
            if 'BASIC QUALIFICATIONS' in str(pack):
                pack = pack.replace('BASIC QUALIFICATIONS\n', '')
            if 'Basic Qualifications:' in str(pack):
                pack = pack.replace('Basic Qualifications:\n', '')
            if '-' in str(pack):
                pack = pack.replace('-', ' ')
            return pack

    def _outtosheet(self, ind, pack, comp):
        c = 0
        for x in range(len(pack)):
            ins = self.sanatize(pack[x])
            if ins:
                if not x == len(pack) - 1:
                    index = '%s%s' % (get_column_letter(ind), x + 1)
                    self.sheet[index] = ins
                    self.sheet[index].font = self.font
                    self.sheet[index].alignment = self.align
                else:
                    ins = ins.split('\n')
                    for y in range(len(ins)):
                        if ins[y] == '':
                            c += 1
                            continue
                        axis = x + 1 + y - c
                        index = '%s%s' % (get_column_letter(ind), axis)
                        try:
                            self.sheet[index] = (ins[y].encode("utf8","backslashreplace")).decode("utf8","backslashreplace")
                        except openpyxl.utils.exceptions.IllegalCharacterError:
                            self.sheet[index] = bs(ins[y], 'lxml').text
                        self.sheet[index].font = self.font
                        self.sheet[index].alignment = self.align

    def compcheck(self, comp):
        if not comp in self.companies:
            self.wb.create_sheet(comp)
            self.sheet = self.wb.get_sheet_by_name(comp)
            self.sheet.page_setup.fitToWidth = 1
            for x in range(len(self.rows) - 1):
                index = '%s%s' % ('A', x + 1)
                self.sheet[index] = self.rows[1:][x] + ':'
                self.sheet[index].font = self.font
                self.sheet[index].alignment = self.align
            self.companies.add(comp)
        self.sheet_switch(comp)

    def sheet_switch(self, name):
        '''
        Switches between active sheets
        '''
        self.sheetn = name
        self.sheet = self.wb.get_sheet_by_name(self.sheetn)

    def update(self, dick):
        '''
        Must be at two story dictionary object with a unique column
        Adds to current dictionary object
        '''
        assert isinstance(dick, dict), "Object is not a dictionary"
        names = self.db[self.sheetn]["Meta"]["column names"]
        for key, value in dick.items():  # (key : value={subkey:subvalue})
            for x in value.keys():  # x=subkey
                if x not in names:  # if subkey not in column name
                    self.add_column(x)  # if not add it
            # insert it into the data base by applying it's key to its value
            self.db[self.sheetn]['Data'][key] = value
            self._update_uid(key)  # then I add the unique ID key
        #  print(self.db[self.sheetn]['Data']) thingy
        k = all([key in self.db[self.sheetn]['Data'].keys()
                 for key in dick.keys()])
        self.lastdick = dick
        if not k:
            self.update(dick)

    def _update_uid(self, uid):
        # Here I get the number of Unique ID's
        leng = len(self.db[self.sheetn]["Data"])
        # assign the new line to a variable
        # add a new key value that is a increment of the UID
        keys = list(self.db[self.sheetn]["Data"].keys())
        for x in range(2, len(self.db[self.sheetn]["Data"]) + 2):
            self.db[self.sheetn]["Meta"]["uid key dic"][keys[x - 2]] = x
        self.db[self.sheetn]["Meta"]['bottom'] = leng + 1

    def add_column(self, name):
        self.db[self.sheetn]["Meta"]["column names"] += [name]
        dick = self.db[self.sheetn]["Meta"]["column dic"]
        vert = column_index_from_string
        try:
            maxi = max([vert(x) for x in dick.keys()]) + 1
        except ValueError:
            maxi = 1
        let = get_column_letter(maxi)
        self.db[self.sheetn]["Meta"]["column dic"][let] = name
        self.db[self.sheetn]["Meta"]["column dic r"] = \
            dict(map(reversed, self.db[self.sheetn]
                     ["Meta"]["column dic"].items()))
        self.db[self.sheetn]["Meta"]["width"] += 1

    def commit_to_xlxs(self):
        # self._commit()# EDITED WARNING WARNING WARNING
        choices = []
        if self.filepath:
            if self.a:
                ans = self.a
            else:
                choices.append("Overwrite?")
                choices.appsend("Create copy")
                ans = utility.get_select(choices)
            if ans in ["Overwrite?", 1]:
                self.wb.save(self.filepath)
            else:
                self.wb.save(self.filepath[:-5] + 1 + '.xlsx')
        else:
            self.wb.save(input('Please enter filename: ') + '.xlsx')

    def _commit(self):
        dick = self.db[self.sheetn]["Meta"]["column dic r"]
        names = self.db[self.sheetn]["Meta"]["column names"]
        data = self.db[self.sheetn]['Data']
        uidic = self.db[self.sheetn]["Meta"]["uid key dic"]
        conv = self.db[self.sheetn]["Meta"]["column dic r"]
        for key in data.keys():
            for k, v in data[key].items():
                if key == v:
                    self.ucolumn = key
                    break
            break

        for name in names:
            try:
                self.sheet['%s%s' % (conv[name], 1)] = name
            except KeyError:
                print("Key error has occured")
                print("%s  %s" % (conv, name))
                exit()
        for key in data.keys():
            for col in names:
                try:
                    self.sheet['%s%s' %
                               (dick[col], uidic[key])] = data[key][col]
                except KeyError:
                    print("a Key error has occurred in _commit")
                    print(key)
                    print(col)
                    self.update(self.lastdick)

        self._sheet_load()

    def self_scan(self):
        db = {x: dict() for x in
              self.db[self.sheetn]["Meta"]["uid key dic"].keys()}
        for x in db.keys():
            for y in self.db[self.sheetn]["Meta"]["column names"]:
                try:
                    db[x][y] =\
                        self.sheet['%s%s' %
                                   (self.db[self.sheetn]["Meta"]["column dic r"][y],
                                    self.db[self.sheetn]["Meta"]["uid key dic"][x])].value
                except:
                    print("an error occured")
                    print(self.db[self.sheetn]["Meta"]["column dic r"])
                    print(self.db[self.sheetn]["Meta"]["column names"])
                    print(y)
        return db

    def _current_line(self):
        '''
        Gives the current bottom line of the excel DB
        '''
        line = 1
        while self.sheet['A%s' % (line)].value is not None:
            line += 1

        return line

    def _current_width(self):
        '''
        Gives the current side line of the excel DB
        '''
        line = 1
        while self.sheet['%s1' % (get_column_letter(line))].value is not None:
            line += 1
        return line

    def _column_scan(self):
        '''
        Gives the full list of column names
        '''
        col = 1
        indexes = []
        while self.sheet['%s1' % get_column_letter(col)].value is not None:
            indexes += [self.sheet['%s1' % get_column_letter(col)].value]
            col += 1
        return indexes

        # Utility

    def scan_list(self, ind):  # Make discrete and change name
        you = []
        cl = self.db[self.sheetn]["Meta"]["bottom"]
        s = 1
        for x in range(s, cl):
            if self.sheet['%s%s' % (ind, x)].value:
                inser = (self.sheet['%s%s' % (ind, x)].value, x)
            else:
                inser = (str(x - 1), x)
            you.append(inser)
        return you

    def _buf(self, string, space):
        buf = int((space - len(string)) / 4)
        sbuf = ' ' * buf
        return sbuf + string + sbuf


'''
    def __str__(self):
        maxs = [len(x) for x in self.column_names]
        for x in range(len(maxs)):
            for y in range(2, self.c_lsline):
                v = self.sheet['%s%s' % (get_column_letter(x + 1), y)].value
                if len(str(v)) > maxs[x]:
                    maxs[x] = len(str(v))
        string = ''.join([self._buf(self.column_names[x], maxs[x])
                  for x in range(len(self.column_names))])
        print (maxs)
        return string
        ###############
        for x in self.db.keys():
            for y in self.db[x].keys():
                iota = self.db[x][y]
                if isinstance(iota, str):
                    if len(iota) > maxlen:
                        maxlen = len(iota)
'''
'''select compn, catn, r.title, r.ID,
        dt, town, 
        state, country_code, r.desc, r.p_desc
        FROM (select compn, catn, q.title, q.ID,
                datetime(post_date, 'unixepoch') as dt, q.desc, q.p_desc
                FROM (select compn, catn, d.title, g.ID, d.desc, d.p_desc
                        FROM (SELECT c.name as compn, cat.name as catn, c.ID
                                FROM (SELECT c.name, c.ID
                                FROM (select * from 
                                company_ID
                                NATURAL JOIN 
                                companies) c
                        LEFT JOIN active_jobs a on a.ID = c.ID) c
                        LEFT JOIN (select * from 
                                        job_categories
                                        NATURAL JOIN 
                                        categories) cat on cat.ID = c.ID) g
                                LEFT JOIN desc as d on d.ID = g.ID) q
                        LEFT JOIN posting_date p on p.ID = q.id) r
                        LEFT JOIN location l on l.ID = r.ID;'''


if __name__ == "__main__":
    app = Excel_bot("/data/bookofjob.db")
    app.initialize()
